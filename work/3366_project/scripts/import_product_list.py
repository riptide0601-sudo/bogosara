"""One-off importer: 제품명/브랜드/전성분 3열짜리 엑셀을 DB에 추가한다.

"화장품 제품 추가.xlsx"처럼 단일 시트에 (제품명, 브랜드, 전성분 원문) 3열이 있는
파일용. 세럼_성분_리스트.xlsx(수분/진정/미백/모공 시트별 concern 태깅)와 달리
concern 태깅 없이 제품·성분 연결만 한다.

행마다:
  - 제품명으로 Product를 찾거나 새로 만든다(브랜드 컬럼도 같이 저장, 카테고리는
    자동 분류).
  - 전성분 원문을 성분 토큰으로 분리해(app.ingredient_list_parsing — 쉼표/골뱅이(@)
    구분자 자동 감지, "*" 표시·각주 제거) 각각 표준 성분과 매칭하고
    product_ingredient로 연결한다(label_rank = 목록 내 순서).

매칭 안 된 토큰은 그냥 연결 안 하고 넘어간다(자유 텍스트로 새 Ingredient를
만들지 않음) — 끝나면 매칭률과 안 된 토큰 목록을 출력한다.

Usage:
    python -m scripts.import_product_list [--db-url URL] [--xlsx-path PATH] [--dry-run]

--dry-run은 아무것도 안 쓰고 매칭률만 보여준다. 재실행해도 안전
(product_ingredient는 ON CONFLICT DO NOTHING, product는 이름 있으면 재사용).
"""

import argparse
from collections import Counter

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.config import settings
from app.ingredient_list_parsing import clean_for_matching, parse_ingredient_tokens
from app.ingredient_matching import match_ingredient_ids
from scripts._product_import_helpers import find_or_create_product, upsert_product_ingredient


def run(db_url: str, xlsx_path: str, dry_run: bool) -> None:
    engine = create_engine(db_url)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    # 시트에 서식만 있고 값은 없는 빈 열(D열 이후)이 따라붙는 파일도 있어 앞 3열만 쓴다.
    rows = [r[:3] for r in list(ws.iter_rows(values_only=True))[1:] if r and r[0] is not None]

    new_products = 0
    existing_products = 0
    total_tokens = 0
    matched_tokens = 0
    unmatched_counter: Counter[str] = Counter()

    with Session(engine) as db:
        for product_name, brand, raw_ingredients in rows:
            product_name = product_name.strip()
            brand = brand.strip() if isinstance(brand, str) and brand.strip() else None

            product, created = find_or_create_product(db, product_name, brand=brand)
            new_products += created
            existing_products += not created

            tokens = parse_ingredient_tokens(raw_ingredients or "")
            for rank, token in enumerate(tokens, start=1):
                total_tokens += 1
                clean = clean_for_matching(token)
                ids = match_ingredient_ids(clean, db)
                if not ids:
                    unmatched_counter[token] += 1
                    continue
                matched_tokens += 1
                if not dry_run:
                    for ingredient_id in ids:
                        upsert_product_ingredient(db, product.product_id, ingredient_id, rank, token)

        if dry_run:
            db.rollback()
        else:
            db.commit()

    mode = "DRY RUN" if dry_run else "APPLIED"
    print(f"[{mode}] db={db_url}")
    print(f"products: {new_products} new, {existing_products} already existed")
    if total_tokens:
        print(f"ingredient tokens: {matched_tokens}/{total_tokens} matched ({matched_tokens / total_tokens:.1%})")
    print(f"top unmatched tokens ({len(unmatched_counter)} unique):")
    for token, count in unmatched_counter.most_common(20):
        print(f"  {count:3d}x  {token}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--db-url", default=settings.database_url)
    parser.add_argument("--xlsx-path", default="/home/jovyan/화장품 제품 추가.xlsx")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    run(args.db_url, args.xlsx_path, args.dry_run)
