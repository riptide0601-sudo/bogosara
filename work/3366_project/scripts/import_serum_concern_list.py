"""One-off importer: load 세럼_성분_리스트.xlsx (수분/진정/미백/모공 시트) into the DB.

Each sheet is a skin concern; each row is (번호, 제품명, 전성분 raw text). For every row:
  - find or create the Product (exact product_name match; new ones get an
    auto-classified category, like POST /products does)
  - parse the raw 전성분 text into ordered ingredient tokens and match each one
    against the ingredient table (same tiered algorithm as OCR/search), linking
    via product_ingredient (label_rank = position in the list)
  - tag the product with this sheet's concern via product_concern

Unmatched tokens are just left unlinked (no new Ingredient rows are invented from
free text) — the script prints an unmatched-token report so gaps are visible.

Usage:
    python -m scripts.import_serum_concern_list [--db-url URL] [--xlsx-path PATH] [--dry-run]

--dry-run parses everything and reports match rates without writing anything.
Safe to re-run — all inserts are ON CONFLICT DO NOTHING / idempotent.
"""

import argparse
from collections import Counter

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.dialects.sqlite import insert as sqlite_insert
from sqlalchemy.orm import Session

from app.config import settings
from app.ingredient_list_parsing import clean_for_matching, parse_ingredient_tokens
from app.ingredient_matching import match_ingredient_ids
from app.models.product_concern import ProductConcern
from scripts._product_import_helpers import find_or_create_product, upsert_product_ingredient


def _upsert_product_concern(db: Session, product_id: str, concern: str) -> None:
    insert_fn = pg_insert if db.bind.dialect.name == "postgresql" else sqlite_insert
    stmt = (
        insert_fn(ProductConcern)
        .values(product_id=product_id, concern=concern)
        .on_conflict_do_nothing(index_elements=["product_id", "concern"])
    )
    db.execute(stmt)


def run(db_url: str, xlsx_path: str, dry_run: bool) -> None:
    engine = create_engine(db_url)
    from app.database import Base

    if not dry_run:
        Base.metadata.create_all(engine, tables=[ProductConcern.__table__])

    wb = openpyxl.load_workbook(xlsx_path)

    new_products = 0
    existing_products = 0
    total_tokens = 0
    matched_tokens = 0
    unmatched_counter: Counter[str] = Counter()

    with Session(engine) as db:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [r for r in list(ws.iter_rows(values_only=True))[1:] if r and r[0] is not None]

            for _, product_name, raw_ingredients in rows:
                product_name = product_name.strip()
                product, created = find_or_create_product(db, product_name)
                new_products += created
                existing_products += not created

                if not dry_run:
                    _upsert_product_concern(db, product.product_id, sheet_name)

                tokens = parse_ingredient_tokens(raw_ingredients or "")
                for rank, token in enumerate(tokens, start=1):
                    total_tokens += 1
                    clean = clean_for_matching(token)
                    ids = match_ingredient_ids(clean, db)
                    if not ids:
                        unmatched_counter[token] += 1
                        continue
                    matched_tokens += 1
                    if not dry_run:
                        # 하나의 토큰에서 여러 성분이 복원된 경우 label_rank가 같은 여러 행으로 들어간다.
                        for ingredient_id in ids:
                            upsert_product_ingredient(db, product.product_id, ingredient_id, rank, token)

        if dry_run:
            db.rollback()
        else:
            db.commit()

    mode = "DRY RUN" if dry_run else "APPLIED"
    print(f"[{mode}] db={db_url}")
    print(f"products: {new_products} new, {existing_products} already existed")
    print(f"ingredient tokens: {matched_tokens}/{total_tokens} matched ({matched_tokens / total_tokens:.1%})")
    print(f"top unmatched tokens ({len(unmatched_counter)} unique):")
    for token, count in unmatched_counter.most_common(20):
        print(f"  {count:3d}x  {token}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--db-url", default=settings.database_url)
    parser.add_argument("--xlsx-path", default="/home/jovyan/세럼_성분_리스트.xlsx")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    run(args.db_url, args.xlsx_path, args.dry_run)
